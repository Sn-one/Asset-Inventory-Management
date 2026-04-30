import io
import json
import math
import os
from datetime import date, datetime, timezone
from urllib.parse import quote_plus

import openpyxl
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.dependencies import get_db
from app.core.security import create_access_token, decode_token, get_password_hash, verify_password
from app.db.models import Asset, AssetMovement, Location, RFIDReader, Role, User, UserRole
from app.schemas.asset import CATEGORIES, CATEGORY_CODES, AssetStatus
from app.services.asset_service import can_transition

_PEMO_UPLOAD_DIR = "app/static/uploads/pemo"
os.makedirs(_PEMO_UPLOAD_DIR, exist_ok=True)

ROLES = ["admin", "viewer", "asset_manager", "maintenance_supervisor", "maintenance_manager"]

ROLE_LABELS = {
    "admin":                  "Administrator",
    "viewer":                 "Viewer",
    "asset_manager":          "Asset Manager",
    "maintenance_supervisor": "Maintenance Supervisor",
    "maintenance_manager":    "Maintenance Manager",
}

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def _get_user(request: Request, db: Session) -> User | None:
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = decode_token(token)
        email = payload.get("sub")
        return db.query(User).filter(User.email == email).first() if email else None
    except ValueError:
        return None


def _require_user(request: Request, db: Session) -> User | RedirectResponse:
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    return user


def _ctx(request: Request, user: User, **kwargs) -> dict:
    role = user.primary_role if user else "viewer"
    return {"request": request, "user": user, "user_role": role, "role_labels": ROLE_LABELS, **kwargs}


def _is_admin(user: User) -> bool:
    return user.primary_role == "admin"


def _can_write(user: User) -> bool:
    return user.primary_role in ("admin", "asset_manager", "maintenance_manager")


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@router.get("/", response_class=HTMLResponse)
def root(request: Request):
    token = request.cookies.get("access_token")
    if token:
        try:
            decode_token(token)
            return RedirectResponse(url="/home", status_code=302)
        except ValueError:
            pass
    return RedirectResponse(url="/login", status_code=302)


@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    token = request.cookies.get("access_token")
    if token:
        try:
            decode_token(token)
            return RedirectResponse(url="/home", status_code=302)
        except ValueError:
            pass
    return templates.TemplateResponse("login.html", {"request": request, "year": datetime.now().year})


@router.post("/login")
def login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.email == email).first()
    if not user or not verify_password(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Invalid email or password", "year": datetime.now().year},
            status_code=401,
        )
    token = create_access_token(user.email)
    response = RedirectResponse(url="/home", status_code=303)
    # secure=True required for HTTPS (Codespace proxy); samesite="none" allows cross-origin forwarding
    response.set_cookie(
        key="access_token", value=token,
        httponly=True, samesite="none", secure=True, max_age=86400,
    )
    return response


@router.post("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("access_token", samesite="none", secure=True)
    return response


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@router.get("/home", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    assets = db.query(Asset).order_by(Asset.id.desc()).all()
    stats = {
        "total":     len(assets),
        "in_stock":  sum(1 for a in assets if a.status == "in_stock"),
        "assigned":  sum(1 for a in assets if a.status == "assigned"),
        "in_repair": sum(1 for a in assets if a.status == "in_repair"),
        "retired":   sum(1 for a in assets if a.status == "retired"),
        "disposed":  sum(1 for a in assets if a.status == "disposed"),
    }
    recent = assets[:5]

    # Build station map data: asset count per site
    asset_counts: dict[int, int] = dict(
        db.query(Asset.site_id, func.count(Asset.id))
        .filter(Asset.site_id.isnot(None))
        .group_by(Asset.site_id)
        .all()
    )
    locations = db.query(Location).filter(
        Location.latitude.isnot(None), Location.longitude.isnot(None)
    ).order_by(Location.name).all()

    locations_json = json.dumps([
        {
            "id": loc.id,
            "name": loc.name,
            "code": loc.code,
            "address": loc.address or "",
            "latitude": loc.latitude,
            "longitude": loc.longitude,
            "asset_count": asset_counts.get(loc.id, 0),
        }
        for loc in locations
    ])

    return templates.TemplateResponse("home.html", _ctx(
        request, user,
        stats=stats,
        recent=recent,
        locations_json=locations_json,
        active_page="home",
    ))


# ---------------------------------------------------------------------------
# Assets — auto-tag generation (must be before /{asset_id} route)
# ---------------------------------------------------------------------------

@router.get("/assets/next-tag")
def asset_next_tag(
    category: str = "",
    site_code: str = "",
    db: Session = Depends(get_db),
):
    cat_code = CATEGORY_CODES.get(category, "OTH")
    site_part = site_code.upper()[:5] if site_code else "GEN"
    prefix = f"{cat_code}-{site_part}-"

    count = db.query(func.count(Asset.id)).filter(Asset.asset_tag.like(f"{prefix}%")).scalar() or 0
    tag = f"{prefix}{count + 1:04d}"

    # Guarantee uniqueness
    while db.query(Asset).filter(Asset.asset_tag == tag).first():
        count += 1
        tag = f"{prefix}{count + 1:04d}"

    return JSONResponse({"tag": tag})


# ---------------------------------------------------------------------------
# Assets — list
# ---------------------------------------------------------------------------

@router.get("/assets", response_class=HTMLResponse)
def assets_list(
    request: Request,
    q: str = "",
    status: str = "",
    category: str = "",
    site: str = "",
    page: int = 1,
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    per_page = 15
    query = db.query(Asset)

    if q:
        like = f"%{q}%"
        query = query.filter(
            Asset.asset_tag.ilike(like)
            | Asset.name.ilike(like)
            | Asset.serial_no.ilike(like)
            | Asset.brand.ilike(like)
            | Asset.rfid_tag.ilike(like)
        )
    if status:
        query = query.filter(Asset.status == status)
    if category:
        query = query.filter(Asset.category == category)
    if site:
        query = query.filter(Asset.site_id == int(site))

    total = query.count()
    pages = max(1, math.ceil(total / per_page))
    page = max(1, min(page, pages))
    assets = query.order_by(Asset.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    locations = db.query(Location).order_by(Location.name).all()

    return templates.TemplateResponse("assets/list.html", _ctx(
        request, user,
        assets=assets,
        total=total,
        page=page,
        pages=pages,
        per_page=per_page,
        q=q,
        status_filter=status,
        category_filter=category,
        site_filter=site,
        categories=CATEGORIES,
        statuses=list(AssetStatus),
        locations=locations,
        active_page="assets",
    ))


# ---------------------------------------------------------------------------
# Assets — create
# ---------------------------------------------------------------------------

@router.get("/assets/new", response_class=HTMLResponse)
def asset_new(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    locations = db.query(Location).order_by(Location.name).all()
    return templates.TemplateResponse("assets/new.html", _ctx(
        request, user,
        categories=CATEGORIES,
        locations=locations,
        active_page="assets",
    ))


@router.post("/assets")
def asset_create(
    request: Request,
    asset_tag: str = Form(...),
    name: str = Form(""),
    brand: str = Form(""),
    model_no: str = Form(""),
    serial_no: str = Form(""),
    category: str = Form(""),
    site_id: str = Form(""),
    rfid_tag: str = Form(""),
    department: str = Form(""),
    assigned_to: str = Form(""),
    purchase_date: str = Form(""),
    purchase_cost: str = Form(""),
    warranty_expiry: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _can_write(user):
        return RedirectResponse(url="/assets?error=Access+denied", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()

    if db.query(Asset).filter(Asset.asset_tag == asset_tag).first():
        return templates.TemplateResponse("assets/new.html", _ctx(
            request, user,
            categories=CATEGORIES,
            locations=locations,
            active_page="assets",
            error=f"Asset tag '{asset_tag}' already exists",
        ), status_code=400)

    _rfid = rfid_tag.strip() or None
    if _rfid and db.query(Asset).filter(Asset.rfid_tag == _rfid).first():
        return templates.TemplateResponse("assets/new.html", _ctx(
            request, user,
            categories=CATEGORIES,
            locations=locations,
            active_page="assets",
            error=f"RFID tag '{_rfid}' is already assigned to another asset",
        ), status_code=400)

    _site_id = int(site_id) if site_id else None
    _location = None
    if _site_id:
        site = db.query(Location).filter(Location.id == _site_id).first()
        _location = site.name if site else None

    def _date(s: str) -> date | None:
        return date.fromisoformat(s) if s else None

    def _float(s: str) -> float | None:
        try:
            return float(s) if s else None
        except ValueError:
            return None

    asset = Asset(
        asset_tag=asset_tag.strip(),
        name=name.strip() or None,
        brand=brand.strip() or None,
        model_no=model_no.strip() or None,
        serial_no=serial_no.strip() or None,
        category=category or None,
        site_id=_site_id,
        location=_location,
        rfid_tag=_rfid,
        department=department.strip() or None,
        assigned_to=assigned_to.strip() or None,
        purchase_date=_date(purchase_date),
        purchase_cost=_float(purchase_cost),
        warranty_expiry=_date(warranty_expiry),
        notes=notes.strip() or None,
        status=AssetStatus.IN_STOCK.value,
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return RedirectResponse(url=f"/assets/{asset.id}?success=Asset+created", status_code=303)


# ---------------------------------------------------------------------------
# Assets — bulk upload via Excel  (must be before /{asset_id})
# ---------------------------------------------------------------------------

@router.get("/assets/bulk-upload", response_class=HTMLResponse)
def assets_bulk_upload_form(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _can_write(user):
        return RedirectResponse(url="/assets?error=Access+denied", status_code=302)
    locations = db.query(Location).order_by(Location.name).all()
    return templates.TemplateResponse("assets/bulk_upload.html", _ctx(
        request, user, locations=locations, categories=CATEGORIES, active_page="assets",
    ))


@router.get("/assets/bulk-upload/template")
def assets_bulk_template_dl(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"
    ws.append([
        "asset_tag", "name", "brand", "model_no", "serial_no",
        "category", "status", "department", "location", "assigned_to",
        "purchase_date", "purchase_cost", "warranty_expiry", "rfid_tag", "notes",
    ])
    ws.append([
        "FD-ACC-9001", "Test Dispenser", "Gilbarco", "Model X", "SN12345",
        "Fuel Dispenser", "in_service", "Operations", "Accra Central Station", "",
        "2024-01-15", "15000", "2027-01-15", "RFID-ACC-TEST", "Sample asset",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets_template.xlsx"},
    )


@router.post("/assets/bulk-upload")
async def assets_bulk_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _can_write(user):
        return RedirectResponse(url="/assets?error=Access+denied", status_code=302)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        locations = db.query(Location).order_by(Location.name).all()
        return templates.TemplateResponse("assets/bulk_upload.html", _ctx(
            request, user, locations=locations, categories=CATEGORIES, active_page="assets",
            error="Could not read Excel file. Please use the provided template.",
        ), status_code=400)

    loc_map = {loc.name.lower(): loc.id for loc in db.query(Location).all()}
    valid_statuses = {s.value for s in AssetStatus}

    added, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        (asset_tag, name, brand, model_no, serial_no,
         category, status, department, location, assigned_to,
         purchase_date, purchase_cost, warranty_expiry, rfid_tag, notes) = (
            row[j] if j < len(row) else None for j in range(15)
        )
        if not asset_tag or not name:
            errors.append(f"Row {i}: asset_tag and name are required")
            continue
        if db.query(Asset).filter(Asset.asset_tag == str(asset_tag)).first():
            skipped += 1
            continue
        status_val = str(status).strip() if status else "in_stock"
        if status_val not in valid_statuses:
            status_val = "in_stock"
        site_id = loc_map.get(str(location).strip().lower()) if location else None

        def _parse_date(v):
            if not v:
                return None
            try:
                return date.fromisoformat(str(v)[:10])
            except Exception:
                return None

        try:
            db.add(Asset(
                asset_tag=str(asset_tag).strip(),
                name=str(name).strip(),
                brand=str(brand).strip() if brand else None,
                model_no=str(model_no).strip() if model_no else None,
                serial_no=str(serial_no).strip() if serial_no else None,
                category=str(category).strip() if category else None,
                status=status_val,
                department=str(department).strip() if department else None,
                location=str(location).strip() if location else None,
                site_id=site_id,
                assigned_to=str(assigned_to).strip() if assigned_to else None,
                purchase_date=_parse_date(purchase_date),
                purchase_cost=float(purchase_cost) if purchase_cost else None,
                warranty_expiry=_parse_date(warranty_expiry),
                rfid_tag=str(rfid_tag).strip() if rfid_tag else None,
                notes=str(notes).strip() if notes else None,
            ))
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    msg = f"{added} asset(s) imported, {skipped} skipped (duplicate tags)"
    if errors:
        msg += f". Errors: {'; '.join(errors[:3])}"
    return RedirectResponse(url=f"/assets?success={quote_plus(msg)}", status_code=303)


# ---------------------------------------------------------------------------
# Assets — detail
# ---------------------------------------------------------------------------

@router.get("/assets/{asset_id}", response_class=HTMLResponse)
def asset_detail(asset_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets?error=Asset+not+found", status_code=302)

    current = AssetStatus(asset.status)
    available = [s for s in AssetStatus if s != current and can_transition(current, s)]

    return templates.TemplateResponse("assets/detail.html", _ctx(
        request, user,
        asset=asset,
        available_statuses=available,
        success=request.query_params.get("success"),
        error=request.query_params.get("error"),
        active_page="assets",
    ))


# ---------------------------------------------------------------------------
# Assets — edit
# ---------------------------------------------------------------------------

@router.get("/assets/{asset_id}/edit", response_class=HTMLResponse)
def asset_edit(asset_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()
    return templates.TemplateResponse("assets/edit.html", _ctx(
        request, user,
        asset=asset,
        categories=CATEGORIES,
        locations=locations,
        active_page="assets",
    ))


@router.post("/assets/{asset_id}/update")
def asset_update(
    asset_id: int,
    request: Request,
    name: str = Form(""),
    brand: str = Form(""),
    model_no: str = Form(""),
    serial_no: str = Form(""),
    category: str = Form(""),
    site_id: str = Form(""),
    rfid_tag: str = Form(""),
    department: str = Form(""),
    assigned_to: str = Form(""),
    purchase_date: str = Form(""),
    purchase_cost: str = Form(""),
    warranty_expiry: str = Form(""),
    notes: str = Form(""),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets", status_code=302)

    _rfid = rfid_tag.strip() or None
    if _rfid and _rfid != asset.rfid_tag:
        conflict = db.query(Asset).filter(Asset.rfid_tag == _rfid, Asset.id != asset_id).first()
        if conflict:
            locations = db.query(Location).order_by(Location.name).all()
            return templates.TemplateResponse("assets/edit.html", _ctx(
                request, user,
                asset=asset,
                categories=CATEGORIES,
                locations=locations,
                active_page="assets",
                error=f"RFID tag '{_rfid}' is already assigned to {conflict.asset_tag}",
            ), status_code=400)

    _site_id = int(site_id) if site_id else None
    _location = asset.location
    if _site_id:
        site = db.query(Location).filter(Location.id == _site_id).first()
        _location = site.name if site else None
    elif not _site_id:
        _location = None

    def _date(s: str) -> date | None:
        return date.fromisoformat(s) if s else None

    def _float(s: str) -> float | None:
        try:
            return float(s) if s else None
        except ValueError:
            return None

    asset.name = name.strip() or None
    asset.brand = brand.strip() or None
    asset.model_no = model_no.strip() or None
    asset.serial_no = serial_no.strip() or None
    asset.category = category or None
    asset.site_id = _site_id
    asset.location = _location
    asset.rfid_tag = _rfid
    asset.department = department.strip() or None
    asset.assigned_to = assigned_to.strip() or None
    asset.purchase_date = _date(purchase_date)
    asset.purchase_cost = _float(purchase_cost)
    asset.warranty_expiry = _date(warranty_expiry)
    asset.notes = notes.strip() or None

    db.commit()
    return RedirectResponse(url=f"/assets/{asset_id}?success=Asset+updated", status_code=303)


# ---------------------------------------------------------------------------
# Assets — status update
# ---------------------------------------------------------------------------

@router.post("/assets/{asset_id}/status")
def asset_status(
    asset_id: int,
    request: Request,
    status: str = Form(...),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets?error=Asset+not+found", status_code=303)

    try:
        new_status = AssetStatus(status)
    except ValueError:
        return RedirectResponse(url=f"/assets/{asset_id}?error=Invalid+status", status_code=303)

    if not can_transition(AssetStatus(asset.status), new_status):
        return RedirectResponse(
            url=f"/assets/{asset_id}?error=Cannot+transition+from+{asset.status}+to+{status}",
            status_code=303,
        )

    asset.status = new_status.value
    db.commit()
    return RedirectResponse(
        url=f"/assets/{asset_id}?success=Status+updated+to+{status.replace('_', '+')}",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# Assets — RFID scan (simulate reader confirming asset location)
# ---------------------------------------------------------------------------

@router.post("/assets/{asset_id}/rfid-scan")
def asset_rfid_scan(
    asset_id: int,
    request: Request,
    site_id: str = Form(...),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets?error=Asset+not+found", status_code=303)

    _site_id = int(site_id) if site_id else None
    asset.rfid_last_seen = datetime.now(timezone.utc).replace(tzinfo=None)
    asset.rfid_confirmed_site_id = _site_id

    if _site_id:
        site = db.query(Location).filter(Location.id == _site_id).first()
        if site:
            asset.location = site.name
            asset.site_id = _site_id

    db.commit()
    return RedirectResponse(
        url=f"/assets/{asset_id}?success=RFID+scan+recorded",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# Assets — delete
# ---------------------------------------------------------------------------

@router.post("/assets/{asset_id}/delete")
def asset_delete(asset_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _can_write(user):
        return RedirectResponse(url="/assets?error=Access+denied", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if asset:
        db.delete(asset)
        db.commit()

    return RedirectResponse(url="/assets?success=Asset+deleted", status_code=303)


# ---------------------------------------------------------------------------
# Users — admin management
# ---------------------------------------------------------------------------

@router.get("/users", response_class=HTMLResponse)
def users_list(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/home?error=Access+denied", status_code=302)

    users = db.query(User).order_by(User.full_name).all()
    locations = db.query(Location).order_by(Location.name).all()
    return templates.TemplateResponse("users/list.html", _ctx(
        request, user,
        users=users,
        locations=locations,
        roles=ROLES,
        active_page="users",
        success=request.query_params.get("success"),
        error=request.query_params.get("error"),
    ))


@router.get("/users/new", response_class=HTMLResponse)
def users_new(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/home?error=Access+denied", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()
    return templates.TemplateResponse("users/new.html", _ctx(
        request, user,
        locations=locations,
        roles=ROLES,
        active_page="users",
    ))


@router.post("/users")
def users_create(
    request: Request,
    full_name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    site_id: str = Form(""),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/home?error=Access+denied", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()

    if db.query(User).filter(User.email == email.strip().lower()).first():
        return templates.TemplateResponse("users/new.html", _ctx(
            request, user,
            locations=locations, roles=ROLES, active_page="users",
            error=f"Email '{email}' is already registered",
        ), status_code=400)

    if role not in ROLES:
        return templates.TemplateResponse("users/new.html", _ctx(
            request, user,
            locations=locations, roles=ROLES, active_page="users",
            error="Invalid role selected",
        ), status_code=400)

    _site_id = int(site_id) if site_id else None
    role_obj = db.query(Role).filter(Role.name == role).first()

    new_user = User(
        email=email.strip().lower(),
        full_name=full_name.strip(),
        password_hash=get_password_hash(password),
        is_active=True,
        site_id=_site_id,
    )
    db.add(new_user)
    db.flush()
    if role_obj:
        db.add(UserRole(user_id=new_user.id, role_id=role_obj.id))
    db.commit()

    return RedirectResponse(
        url=f"/users?success={quote_plus(full_name.strip() + ' created successfully')}",
        status_code=303,
    )


@router.post("/users/{user_id}/toggle")
def users_toggle(user_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/home?error=Access+denied", status_code=302)

    target = db.query(User).filter(User.id == user_id).first()
    if target and target.id != user.id:
        target.is_active = not target.is_active
        db.commit()

    action = "activated" if (target and target.is_active) else "deactivated"
    return RedirectResponse(url=f"/users?success=User+{action}", status_code=303)


# ---------------------------------------------------------------------------
# Sites
# ---------------------------------------------------------------------------

@router.get("/sites", response_class=HTMLResponse)
def sites_list(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()

    asset_counts: dict[int, int] = dict(
        db.query(Asset.site_id, func.count(Asset.id))
        .filter(Asset.site_id.isnot(None))
        .group_by(Asset.site_id)
        .all()
    )
    rfid_counts: dict[int, int] = dict(
        db.query(RFIDReader.site_id, func.count(RFIDReader.id))
        .filter(RFIDReader.site_id.isnot(None), RFIDReader.is_active == True)  # noqa: E712
        .group_by(RFIDReader.site_id)
        .all()
    )

    locations_json = json.dumps([
        {
            "id": loc.id,
            "name": loc.name,
            "code": loc.code,
            "address": loc.address or "",
            "latitude": loc.latitude,
            "longitude": loc.longitude,
            "asset_count": asset_counts.get(loc.id, 0),
            "rfid_count": rfid_counts.get(loc.id, 0),
        }
        for loc in locations
        if loc.latitude and loc.longitude
    ])

    return templates.TemplateResponse("sites/list.html", _ctx(
        request, user,
        locations=locations,
        asset_counts=asset_counts,
        rfid_counts=rfid_counts,
        locations_json=locations_json,
        active_page="sites",
        success=request.query_params.get("success"),
        error=request.query_params.get("error"),
    ))


# ---------------------------------------------------------------------------
# RFID Audit
# ---------------------------------------------------------------------------

@router.get("/rfid/audit", response_class=HTMLResponse)
def rfid_audit(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    rfid_assets = (
        db.query(Asset)
        .filter(Asset.rfid_tag.isnot(None))
        .order_by(Asset.rfid_last_seen.desc().nullslast())
        .all()
    )
    mobile_readers = (
        db.query(RFIDReader)
        .filter(RFIDReader.reader_type == "mobile", RFIDReader.is_active == True)  # noqa: E712
        .all()
    )
    locations = db.query(Location).order_by(Location.name).all()

    return templates.TemplateResponse("rfid/audit.html", _ctx(
        request, user,
        rfid_assets=rfid_assets,
        mobile_readers=mobile_readers,
        locations=locations,
        success=request.query_params.get("success"),
        error=request.query_params.get("error"),
        active_page="rfid",
    ))


@router.post("/rfid/scan")
def rfid_scan(
    request: Request,
    rfid_tag: str = Form(...),
    site_id: str = Form(...),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    asset = db.query(Asset).filter(Asset.rfid_tag == rfid_tag.strip()).first()
    if not asset:
        return RedirectResponse(url="/rfid/audit?error=RFID+tag+not+found", status_code=303)

    _site_id = int(site_id) if site_id else None
    asset.rfid_last_seen = datetime.now(timezone.utc).replace(tzinfo=None)
    asset.rfid_confirmed_site_id = _site_id

    if _site_id:
        site = db.query(Location).filter(Location.id == _site_id).first()
        if site:
            asset.location = site.name
            asset.site_id = _site_id

    db.commit()
    msg = quote_plus(f"Scanned {asset.asset_tag} confirmed at {asset.location or 'site'}")
    return RedirectResponse(url=f"/rfid/audit?success={msg}", status_code=303)


# ---------------------------------------------------------------------------
# Assets — Move (PEMO)
# ---------------------------------------------------------------------------

@router.get("/assets/{asset_id}/move", response_class=HTMLResponse)
def asset_move_form(asset_id: int, request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if user.primary_role not in ("admin", "asset_manager"):
        return RedirectResponse(url=f"/assets/{asset_id}?error=Access+denied", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets?error=Asset+not+found", status_code=302)

    locations = db.query(Location).order_by(Location.name).all()
    history = (
        db.query(AssetMovement)
        .filter(AssetMovement.asset_id == asset_id)
        .order_by(AssetMovement.created_at.desc())
        .all()
    )
    return templates.TemplateResponse("assets/move.html", _ctx(
        request, user,
        asset=asset,
        locations=locations,
        history=history,
        active_page="assets",
        error=request.query_params.get("error"),
    ))


@router.post("/assets/{asset_id}/move")
async def asset_move(
    asset_id: int,
    request: Request,
    to_site_id: str = Form(...),
    pemo_number: str = Form(...),
    notes: str = Form(""),
    pemo_document: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if user.primary_role not in ("admin", "asset_manager"):
        return RedirectResponse(url=f"/assets/{asset_id}?error=Access+denied", status_code=302)

    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        return RedirectResponse(url="/assets?error=Asset+not+found", status_code=302)

    if not pemo_number.strip():
        locations = db.query(Location).order_by(Location.name).all()
        return templates.TemplateResponse("assets/move.html", _ctx(
            request, user, asset=asset, locations=locations, history=[], active_page="assets",
            error="PEMO number is required",
        ), status_code=400)

    # Save uploaded PEMO document
    doc_path = None
    if pemo_document and pemo_document.filename:
        safe_name = f"{pemo_number.strip().replace('/', '-')}_{pemo_document.filename}"
        doc_path = os.path.join(_PEMO_UPLOAD_DIR, safe_name)
        content = await pemo_document.read()
        with open(doc_path, "wb") as f:
            f.write(content)
        doc_path = f"uploads/pemo/{safe_name}"

    _to_site_id = int(to_site_id) if to_site_id else None
    from_site_id = asset.site_id

    movement = AssetMovement(
        asset_id=asset.id,
        from_site_id=from_site_id,
        to_site_id=_to_site_id,
        pemo_number=pemo_number.strip(),
        pemo_document_path=doc_path,
        notes=notes.strip() or None,
        moved_by_id=user.id,
    )
    db.add(movement)

    if _to_site_id:
        site = db.query(Location).filter(Location.id == _to_site_id).first()
        if site:
            asset.site_id = _to_site_id
            asset.location = site.name

    db.commit()
    return RedirectResponse(
        url=f"/assets/{asset_id}?success={quote_plus('Asset moved — PEMO ' + pemo_number.strip() + ' recorded')}",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# Sites — add single / bulk upload
# ---------------------------------------------------------------------------

@router.get("/sites/new", response_class=HTMLResponse)
def sites_new(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/sites?error=Access+denied", status_code=302)
    return templates.TemplateResponse("sites/new.html", _ctx(request, user, active_page="sites"))


@router.post("/sites")
def sites_create(
    request: Request,
    name: str = Form(...),
    code: str = Form(...),
    address: str = Form(""),
    latitude: str = Form(""),
    longitude: str = Form(""),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/sites?error=Access+denied", status_code=302)

    code_clean = code.strip().upper()
    if db.query(Location).filter(Location.code == code_clean).first():
        return templates.TemplateResponse("sites/new.html", _ctx(
            request, user, active_page="sites",
            error=f"Station code '{code_clean}' already exists",
        ), status_code=400)

    db.add(Location(
        name=name.strip(),
        code=code_clean,
        address=address.strip() or None,
        latitude=float(latitude) if latitude.strip() else None,
        longitude=float(longitude) if longitude.strip() else None,
    ))
    db.commit()
    return RedirectResponse(url=f"/sites?success={quote_plus(name.strip() + ' added')}", status_code=303)


@router.get("/sites/bulk-upload", response_class=HTMLResponse)
def sites_bulk_upload_form(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/sites?error=Access+denied", status_code=302)
    return templates.TemplateResponse("sites/bulk_upload.html", _ctx(request, user, active_page="sites"))


@router.get("/sites/bulk-upload/template")
def sites_bulk_template(request: Request, db: Session = Depends(get_db)):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stations"
    ws.append(["name", "code", "address", "latitude", "longitude"])
    ws.append(["Example Station", "EXM", "123 Main St, Accra", "5.5502", "-0.2174"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stations_template.xlsx"},
    )


@router.post("/sites/bulk-upload")
async def sites_bulk_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    user = _get_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    if not _is_admin(user):
        return RedirectResponse(url="/sites?error=Access+denied", status_code=302)

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception:
        return templates.TemplateResponse("sites/bulk_upload.html", _ctx(
            request, user, active_page="sites",
            error="Could not read Excel file. Please use the provided template.",
        ), status_code=400)

    added, skipped = 0, 0
    errors = []
    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        name, code, address, lat, lng = (row[j] if j < len(row) else None for j in range(5))
        if not name or not code:
            errors.append(f"Row {i}: name and code are required")
            continue
        code_clean = str(code).strip().upper()
        if db.query(Location).filter(Location.code == code_clean).first():
            skipped += 1
            continue
        try:
            db.add(Location(
                name=str(name).strip(),
                code=code_clean,
                address=str(address).strip() if address else None,
                latitude=float(lat) if lat else None,
                longitude=float(lng) if lng else None,
            ))
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    msg = f"{added} station(s) added, {skipped} skipped (duplicate codes)"
    if errors:
        msg += f". Errors: {'; '.join(errors[:3])}"
    return RedirectResponse(url=f"/sites?success={quote_plus(msg)}", status_code=303)
